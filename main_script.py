import openpyxl
import json
from pathlib import Path
from datetime import datetime

def determine_sheet_name(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet_names = workbook.sheetnames
        print(f"Available sheet names: {sheet_names}")
        return sheet_names[0]  # Assuming the first sheet, adjust as needed
    except Exception as e:
        print(f"Error determining sheet name: {e}")
        return None

def identify_tables(ws):
    tables = []
    table = []
    headers_found = False
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # Check if the row has any non-None values
            if headers_found:
                if all(cell is None for cell in row):  # End of table
                    tables.append(table)
                    table = []
                    headers_found = False
                else:
                    table.append(row)
            else:
                headers_found = True
                table.append(row)
        elif headers_found:
            tables.append(table)
            table = []
            headers_found = False

    if table:  # If there's an unappended table at the end
        tables.append(table)

    return tables

def convert_date_to_words(date):
    """Convert a date string to a more readable format with month and period."""
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
        return date_obj.strftime("%d %b %Y")
    except ValueError:
        return date
        

def evaluate_formula(worksheet, formula):
    """Evaluate a formula in a given worksheet and return its computed value."""
    try:
        cell = worksheet[formula]
        return cell.value
    except Exception as e:
        print(f"Error evaluating formula: {e}")
        return None


def serialize_tables_to_json(tables, worksheet, output_file):
    all_data = {}
    for i, table in enumerate(tables):
        if table:
            columns = table[0]
            data = []
            for row in table[1:]:
                row_data = {}
                for j, cell in enumerate(row):
                    if columns[j] is not None:
                        if isinstance(cell, datetime):
                            row_data[str(columns[j])] = convert_date_to_words(str(cell))
                        elif isinstance(cell, str) and not cell.startswith('='):
                            row_data[str(columns[j])] = cell
                        elif isinstance(cell, str) and cell.startswith('='):
                            # Skip formula cells
                            continue
                        else:
                            evaluated_value = cell
                            row_data[str(columns[j])] = evaluated_value
                data.append(row_data)
            all_data[f"table_{i+1}"] = data

    try:
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump(all_data, json_file, ensure_ascii=False, indent=4)
        print(f"Data has been successfully extracted to {output_file}")
    except ValueError as e:
        print(f"Error converting tables to JSON: {e}")


def clean_data(ws):
    # Drop rows where all elements are NaN
    cleaned_rows = [row for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row)]
    
    # Ensure column names are strings and strip spaces
    cleaned_data = []
    for row in cleaned_rows:
        cleaned_data.append([str(cell).strip() if cell is not None else cell for cell in row])

    print("Data cleaned.")
    return cleaned_data

def main(input_excel_file, sheet_name, output_json_file):
    # Get the workbook and worksheet
    workbook = openpyxl.load_workbook(input_excel_file)
    worksheet = workbook[sheet_name]

    # Clean the data
    cleaned_data = clean_data(worksheet)
    if not cleaned_data:
        print("Cleaned data is empty. Exiting.")
        return

    # Identify tables
    tables = identify_tables(worksheet)

    # Serialize tables to JSON
    serialize_tables_to_json(tables, worksheet, output_json_file)

if __name__ == "__main__":
    input_excel_file = Path("example_0.xlsx")
    sheet_name = 'Analysis Output'
    output_json_file = 'output.json'

    main(input_excel_file, sheet_name, output_json_file)